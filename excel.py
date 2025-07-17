from openpyxl import Workbook, load_workbook
import os

def iniciar_excel():
    caminho = "monitores_kabum.xlsx"

    if os.path.exists(caminho):
        wb = load_workbook(caminho)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nome", "Preço", "Desconto", "Avaliação", "Frete Grátis"])
        wb.save(caminho)

    return ws, wb, caminho

def salvar_no_excel(ws, wb, caminho, produto):
    ws.append([produto["nome"], produto["preco"], produto["desconto"], produto["avaliacao"], produto["frete"]])
    wb.save(caminho)
    print(f"💾 Produto salvo no Excel: {produto['nome']}")
