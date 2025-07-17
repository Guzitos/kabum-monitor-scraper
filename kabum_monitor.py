import mysql.connector
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
import time
import os
from enviargmail import enviar_email

# Conexão com MySQL
conexao = mysql.connector.connect(
    host="localhost",
    user="seu_usuario",
    password="sua_senha",
    database="seu_bancoDeDados"
)
cursor = conexao.cursor()

# Iniciar o navegador
navegador = webdriver.Chrome()
navegador.maximize_window()

# Configurar Excel
excel_path = "monitores_kabum.xlsx"

if os.path.exists(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Nome", "Preço", "Desconto", "Avaliação", "Frete Grátis"])

    # Variáveis de controle
    pagina = 1
    produtos_coletados = 0
    limite = 5  # Quantos produtos deseja coletar

    while produtos_coletados < limite:
        url_pagina = f"https://www.kabum.com.br/computadores/monitores?page={pagina}"
        navegador.get(url_pagina)
        time.sleep(2)

        produtos = navegador.find_elements(By.CLASS_NAME, "productCard")

        if not produtos:
            print("❌ Nenhum produto encontrado. Encerrando.")
            break

        for produto in produtos:
            if produtos_coletados >= limite:
                break

            nome = produto.find_element(By.CLASS_NAME, "sc-27518a44-8").text
            preco = produto.find_element(By.CLASS_NAME, "sc-57f0fd6e-2").text

            try:
                avaliacao = produto.find_element(By.CLASS_NAME, "leading-none").text
            except:
                avaliacao = "(0)"

            try:
                frete = produto.find_element(By.CLASS_NAME, "w-max").text
            except:
                frete = "Sem frete grátis"

            try:
                desconto = produto.find_element(By.CLASS_NAME, "ary-500").text
            except:
                desconto = "Sem desconto"

            # Verificar se já existe no banco
            cursor.execute("SELECT id FROM monitores WHERE nome = %s", (nome,))
            cursor.fetchall()  # Consome o resultado para evitar erro
            if cursor.rowcount > 0:
                print(f"⚠️ Produto já existe: {nome}")
            else:
                # Inserir no banco
                sql = "INSERT INTO monitores (nome, preco, desconto, avaliacao, frete_gratis) VALUES (%s, %s, %s, %s, %s)"
                valores = (nome, preco, desconto, avaliacao, frete)
                cursor.execute(sql, valores)
                conexao.commit()

                # Inserir no Excel
                ws.append([nome, preco, desconto, avaliacao, frete])
                wb.save(excel_path)

                produtos_coletados += 1
                print(f"✅ Inserido: {produtos_coletados} - {nome}")

        pagina += 1

# Finalização
navegador.quit()
cursor.close()
conexao.close()

enviar_email()
print("✅ Raspagem finalizada com sucesso.")
print(f"📁 Dados salvos em: {excel_path}")
